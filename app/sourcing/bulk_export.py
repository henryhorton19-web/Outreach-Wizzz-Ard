"""Generic bulk-export ingestion and screening (Stage F of EXECUTION_PLAN_5).

Parses tabular exports (CSV, XLSX, JSON) and evaluates them against configured
local gates (revenue_band_min/max, require_keyword_in_field, reject_last_event_types,
exclusion set).
"""
from __future__ import annotations

import csv
import io
import json
from pathlib import Path


def _parse_csv(data: bytes) -> list[dict]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    return [row for row in reader]


def _parse_json(data: bytes) -> list[dict]:
    raw = json.loads(data.decode("utf-8", errors="replace"))
    if isinstance(raw, list):
        return raw
    if isinstance(raw, dict):
        # common API shapes: {"results": [...]} or {"companies": [...]}
        for key in ("results", "companies", "data", "items"):
            if isinstance(raw.get(key), list):
                return raw[key]
    return []


def _parse_xlsx(data: bytes) -> list[dict]:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h or "").strip() for h in rows[0]]
        return [dict(zip(headers, [str(v or "") for v in row])) for row in rows[1:]]
    except ImportError:
        raise RuntimeError("openpyxl is required for XLSX parsing")


def parse_export(data: bytes, filename: str = "") -> list[dict]:
    """Parse bytes from a bulk export into a list of row dicts."""
    fn = (filename or "").lower()
    if fn.endswith(".xlsx"):
        return _parse_xlsx(data)
    if fn.endswith(".json"):
        return _parse_json(data)
    # Default: CSV
    return _parse_csv(data)


def evaluate_local_gates(rows: list[dict], gates: dict) -> tuple[list[dict], list[dict]]:
    """Apply local screening gates to parsed export rows.

    gates keys (all optional):
      - revenue_band_min (int): minimum revenue in USD
      - revenue_band_max (int): maximum revenue in USD
      - require_keyword_in_field (dict[str, str]): {field: keyword} all must match
      - reject_last_event_types (list[str]): reject if last_event_type matches

    Returns (passed, rejected) lists.
    """
    passed, rejected = [], []
    rev_min = gates.get("revenue_band_min")
    rev_max = gates.get("revenue_band_max")
    require_kw = gates.get("require_keyword_in_field") or {}
    reject_event_types = set(gates.get("reject_last_event_types") or [])

    for row in rows:
        reason = None

        # Revenue band filter
        if rev_min is not None or rev_max is not None:
            rev_raw = row.get("revenue") or row.get("annual_revenue") or row.get("revenue_usd") or "0"
            try:
                rev = float(str(rev_raw).replace(",", "").replace("$", "")) if rev_raw else 0
            except (ValueError, TypeError):
                rev = 0
            if rev_min is not None and rev < rev_min:
                reason = f"revenue {rev} < {rev_min}"
            elif rev_max is not None and rev > rev_max:
                reason = f"revenue {rev} > {rev_max}"

        # Required keyword in field
        if not reason:
            for field, keyword in require_kw.items():
                field_val = str(row.get(field) or "").lower()
                if keyword.lower() not in field_val:
                    reason = f"'{keyword}' not found in field '{field}'"
                    break

        # Reject by last event type
        if not reason and reject_event_types:
            last_event = str(row.get("last_event_type") or "").lower()
            if last_event in {t.lower() for t in reject_event_types}:
                reason = f"rejected event type: {last_event}"

        if reason:
            rejected.append({**row, "__reject_reason": reason})
        else:
            passed.append(row)

    return passed, rejected
