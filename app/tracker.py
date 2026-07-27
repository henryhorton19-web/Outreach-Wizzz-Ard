"""Outreach tracker integration.

The Outreach_Tracker.xlsx workbook is the system of record. On ingest we READ target names
(and any known contact/category) from the target-list sheets; on approve we WRITE the sent row
back to the 'Reach Out To' sheet so the existing follow-up workflow stays intact.

This module is deliberately defensive: a missing workbook, an unexpected column layout, or a
locked file must never crash a draft or an approval. Reads return []; writes return a status
dict and swallow their own errors.
"""
from __future__ import annotations

import datetime
import io
from pathlib import Path
from typing import Any

# Sheets we scan for target names, and the column header we treat as the company name.
_TARGET_SHEETS = ("Initial", "Warm Contacts", "Priority")
_NAME_HEADERS = ("Company Name", "Company", "Name")
_CONTACT_NAME_HEADERS = ("Contact Person", "CEO / Target Name", "Contact Name")
_CONTACT_EMAIL_HEADERS = ("Contact Email", "Target Email", "Email Address", "Email")
_CATEGORY_HEADERS = ("Category",)

_REACH_SHEET = "Reach Out To"
_REACH_HEADERS = [
    "Company Name", "Contact Name", "Email Address", "Subject Line", "Send Date",
    "Outreach Status", "Follow-up 1 Date", "Follow-up 1 Status",
    "Follow-up 2 Date", "Follow-up 2 Status", "Official Portal Applied? (Yes/No/NA)",
]


def _header_index(header_row: list[Any], candidates) -> int:
    norm = [str(c).strip().lower() if c is not None else "" for c in header_row]
    for cand in candidates:
        if cand.strip().lower() in norm:
            return norm.index(cand.strip().lower())
    return -1


def parse_tracker_bytes(data: bytes) -> list[dict]:
    """Return [{name, ref?, contact_name?, contact_email?}] read from the target sheets."""
    try:
        import openpyxl
    except Exception:
        return []
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        return []

    rows: list[dict] = []
    seen: set[str] = set()
    for sheet_name in _TARGET_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        it = ws.iter_rows(values_only=True)
        try:
            header = list(next(it))
        except StopIteration:
            continue
        name_i = _header_index(header, _NAME_HEADERS)
        if name_i == -1:
            continue
        cat_i = _header_index(header, _CATEGORY_HEADERS)
        cn_i = _header_index(header, _CONTACT_NAME_HEADERS)
        ce_i = _header_index(header, _CONTACT_EMAIL_HEADERS)
        for r in it:
            if not r or name_i >= len(r):
                continue
            name = r[name_i]
            if not name or not str(name).strip():
                continue
            name = str(name).strip()
            key = name.lower()
            if key in seen:
                continue
            seen.add(key)
            rec: dict = {"name": name}
            if cat_i != -1 and cat_i < len(r) and r[cat_i]:
                rec["ref"] = str(r[cat_i]).strip()
            if cn_i != -1 and cn_i < len(r) and r[cn_i]:
                rec["contact_name"] = str(r[cn_i]).strip()
            if ce_i != -1 and ce_i < len(r) and r[ce_i]:
                rec["contact_email"] = str(r[ce_i]).strip()
            rows.append(rec)
    return rows


def write_reach_row(path: Path, *, company: str, contact_name: str, email: str,
                    subject: str) -> dict:
    """Append/update a row on the 'Reach Out To' sheet. Returns a status dict; never raises."""
    try:
        import openpyxl
    except Exception:
        return {"ok": False, "reason": "openpyxl unavailable"}
    if not path or not Path(path).exists():
        return {"ok": False, "reason": "tracker not found"}
    try:
        wb = openpyxl.load_workbook(path)
        if _REACH_SHEET in wb.sheetnames:
            ws = wb[_REACH_SHEET]
        else:
            ws = wb.create_sheet(_REACH_SHEET)
            ws.append(_REACH_HEADERS)

        # ensure a header row exists
        first = [c.value for c in ws[1]] if ws.max_row >= 1 else []
        if not any(first):
            for i, h in enumerate(_REACH_HEADERS, start=1):
                ws.cell(row=1, column=i, value=h)

        today = datetime.date.today().isoformat()
        # look for an existing row for this company (col 1)
        target_row = None
        for row in ws.iter_rows(min_row=2):
            if row and row[0].value and str(row[0].value).strip().lower() == company.strip().lower():
                target_row = row[0].row
                break
        values = [company, contact_name or "", email or "", subject or "", today,
                  "Sent", "", "", "", "", ""]
        if target_row is None:
            ws.append(values)
        else:
            for i, v in enumerate(values, start=1):
                if v != "":
                    ws.cell(row=target_row, column=i, value=v)
        wb.save(path)
        return {"ok": True, "sheet": _REACH_SHEET}
    except PermissionError:
        return {"ok": False, "reason": "tracker file is open/locked — close it and retry"}
    except Exception as e:
        return {"ok": False, "reason": f"{type(e).__name__}: {e}"}
